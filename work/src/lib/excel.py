import openpyxl
import openpyxl.styles

from lib.log import LOG

class Excel():
    WORKBOOK: openpyxl.workbook.Workbook = None
    WORKSHEET: openpyxl.worksheet.worksheet.Worksheet = None    
    COLUMN_NAMES = {  # 列: 値
        "A": "ステータス",
        "B": "棚卸結果",
        "C": "備考",
        "D": "備考（前回以前）",
        "E": "管理部門",
        "F": "管理番号",
        "G": "シリアル（参考）",
        "H": "稟議番号",
        "I": "管理者",
        "J": "使用場所",
        "K": "使用者"
    }
    COLUMN_NAME_ROW = 2  # 2行目
    START_LOW = 3  # 表の値は3行目から
    LAST_LOW: int = None

    def load(
            self,
            file_path: str,
            sheet_name: str
            ) -> bool:
        """
        Excelファイルを読み込みます。

        Args:
            file_path (str): ファイルパス
            sheet_name (str): 自動入力するシートの名前

        Returns:
            bool: Excelファイルの読み込みに成功したらTrue、失敗したらFalse
        """

        # Excelファイルを開いて対象のシートを読み込みます。
        try:
            self.WORKBOOK = openpyxl.load_workbook(file_path)
            LOG.debug(f"Worksheets: {self.WORKBOOK.sheetnames}")

            if sheet_name in self.WORKBOOK.sheetnames:
                self.WORKSHEET = self.WORKBOOK[sheet_name]
                # 後で上書きするためにWORKBOOKをここでクローズしない
                return True
            else:
                LOG.error(f"The '{sheet_name}' sheet doesn't exist.")
                self.WORKBOOK.close()
                return False
        except FileNotFoundError:
            LOG.error(f"'{file_path}' doesn't exist.")
            return False
        except Exception:
            LOG.exception("Unexpected error occurred.")
            try:
                self.WORKBOOK.close()
            except:
                pass
            return False

    def __are_column_names_valid(
            self
            ) -> bool:
        """
        列名のバリデーションチェックを行います。

        Returns:
            bool: 一致していればTrue、一致していなければFalse
        """

        result = True
        for column_letter, expected_value in self.COLUMN_NAMES.items():
            cell = column_letter + str(self.COLUMN_NAME_ROW)
            actual_value = self.WORKSHEET[cell].value
            if actual_value != expected_value:
                LOG.error(f"The {cell} value was expected to be '{expected_value}', "
                          f"but it was '{actual_value}'.")
                result = False

        return result

    def __find_last_row(
            self
            ) -> int:
        """
        ワークシートから表の最終行の行番号を取得します。
        最終行を見つけることができなかった場合は-1を返します。

        Returns:
            int: 最終行の行番号
        """

        # ステータス列の値は、'棚卸対象'または'対象外'のどちらかの値が入る想定です。
        # 3行目から何行目まで値が入っているかを確認して、表の最終行を見つけます。
        last_row = -1
        for row_num in range(self.START_LOW, self.WORKSHEET.max_row + 1):
            cell = f"A{row_num}"
            if not self.WORKSHEET[cell].value in ["棚卸対象", "対象外"]:
                return last_row
            last_row = row_num

        return -1

    def is_worksheet_vaild(
            self
            ) -> bool:
        """
        ワークシートの整合性を確認します。

        Returns:
            bool: 整合性がある場合はTrue、無い場合はFalseを返します。
        """

        if not self.__are_column_names_valid():
            return False
        
        last_low = self.__find_last_row()
        if last_low != -1:
            self.LAST_LOW = last_low
            return True
        else:
            LOG.error("Failed to find last row of the table from the worksheet.")
            return False

    def load_inventory_data(
            self
            ) -> dict[str, dict[str, str]]:
        """
        棚卸リストを取得します。

        Returns:
            dict[str, dict[str, str]]: 棚卸リストの表データ
        
        Raises:
            Exception: 想定外のエラー
        """

        inventory_data = {}
        try:
            for row_num in range(self.START_LOW, self.LAST_LOW + 1):
                row_data = {}
                for column_letter, column_name in self.COLUMN_NAMES.items():
                    value = self.WORKSHEET[column_letter + str(row_num)].value
                    if value == None:
                        value = ""  # 技術資産管理表の空値に合わせます。
                    row_data[column_name] = str(value)  # str型でない場合があるためstr型にキャストします。
                inventory_data[str(row_num)] = row_data
        except Exception as ex:
            raise ex

        return inventory_data

    def overwrite(
            self,
            diff: dict[str, dict[str, str]],
            file_path: str
            ) -> None:
        """
        Excelファイルをdiffの内容で上書きし、更新されたセルのフォントを赤色にします。

        Args:
            diff (dict[str, dict[str, str]]): Excelと技術資産管理表の差分
            output_dir (str): 出力先フォルダパス
        """

        if self.WORKBOOK is None or self.WORKSHEET is None:
            LOG.error("Workbook or worksheet is not loaded.")
            return

        has_error = False
        for row_num, changes in diff.items():
            for column_name, change in changes.items():
                # COLUMN_NAMESから列の文字を取得
                column_letter = next((k for k, v in self.COLUMN_NAMES.items() if v == column_name), None)
                if column_letter is None:
                    LOG.error(f"Column name '{column_name}' not found in COLUMN_NAMES.")
                    has_error = True
                    continue

                # セルのアドレスを作成
                cell_address = f"{column_letter}{row_num}"
                # セルの値を更新
                self.WORKSHEET[cell_address].value = change["After"]
                if column_name != "棚卸結果":
                    # フォントの色を赤に設定
                    self.WORKSHEET[cell_address].font = openpyxl.styles.Font(color="FF0000")

        if has_error:
            LOG.error("Excel file has not been updated.")
            return
        
        # 変更を保存
        self.WORKBOOK.save(file_path)
        LOG.info(f"Excel file has been updated as '{file_path}'.")
        